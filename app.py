# -*- coding: utf-8 -*-
from datetime import datetime

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pandas import json_normalize

from config import settings

HELP_DICT = {
    '/v2/product/list': None,
    # Вписать нужные product_id через , между []
    '/v2/product/info/list': {'product_id': [326670334], 'offer_id': [], 'sku': []},
    # Вписать нужный product_id
    '/v1/product/info/description': {'product_id': 326670334, 'offer_id': ''}
}
HEADERS = {
    'Client-Id': settings.client_id,
    'Api-Key': settings.api_key,
    'Content-Type': 'application/json'
}

ITEMS_METHODS = ['/v2/product/list', '/v2/product/info/list', ]

if __name__ == '__main__':
    method = '/v2/product/list'
    body = HELP_DICT[method] if HELP_DICT[method] else None
    r = requests.post(f'https://{settings.host}{method}', headers=HEADERS, json=body)
    if method in ITEMS_METHODS:
        result = r.json()['result']['items']
    else:
        result = r.json()['result']
    print(result)

    try:
        # Конвертация результата в Excel
        df = json_normalize(result)
        file_name = f'tables/{method.replace("/", "_")}_{datetime.today().strftime("%d_%m_%Y")}.xlsx'
        df.to_excel(file_name)
        print('Successful conversion')

        # Увеличение ширины колонок
        wb = load_workbook(file_name)
        ws = wb.active
        # для строк
        for i in range(1, ws.max_row + 1):
            # если высота строки не изменялась программно
            # или вручную то `rh` будет присваиваться `None`
            rh = ws.row_dimensions[i].height
            # по умолчанию высота строки равна 15 единицам
            row_heights = 15 if rh is None else rh
            # print(f'Строка {i} имеет высоту {row_heights}')

        # для колонок
        for i in range(1, ws.max_column + 1):
            # преобразовываем индекс столбца в его букву
            letter = get_column_letter(i)
            # получаем ширину столбца
            col_width = ws.column_dimensions[letter].width
            # print(f'Столбец {letter} имеет ширину {col_width}')

        print('Successful cell formatting')
    except Exception as err:
        print(f'Get an Error: {err}')

