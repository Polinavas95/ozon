import openpyxl


def width_column_product_list(file_name):
    worksheet = openpyxl.load_workbook(file_name)
    sheet = worksheet.active
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    worksheet.save(file_name)
    return 'Width changed'


def width_column_product_info_list(file_name):
    worksheet = openpyxl.load_workbook(file_name)
    sheet = worksheet.active
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['H'].width = 25
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['N'].width = 15
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['P'].width = 15
    sheet.column_dimensions['W'].width = 15
    sheet.column_dimensions['Y'].width = 15
    sheet.column_dimensions['AD'].width = 15
    sheet.column_dimensions['AF'].width = 15
    sheet.column_dimensions['AL'].width = 20
    sheet.column_dimensions['AM'].width = 25
    worksheet.save(file_name)
    return 'Width changed'


def width_column_product_description(file_name):
    worksheet = openpyxl.load_workbook(file_name)
    sheet = worksheet.active
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 50
    sheet.column_dimensions['E'].width = 40
    worksheet.save(file_name)
    return 'Width changed'